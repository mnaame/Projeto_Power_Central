from __future__ import annotations

from datetime import datetime, time, timedelta
from pathlib import Path

from flask import (
    Blueprint,
    abort,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required
from openpyxl import load_workbook

from app.domain.dates import FUSO_HORARIO
from app.extensions import db
from app.models.report import ReportRun
from app.services import audit_service, report_service

bp = Blueprint("reports", __name__, url_prefix="/relatorios")

LINHAS_POR_PAGINA = 25
HISTORICO_LIMITE = 20

MODULOS = {
    "atendimentos": {
        "titulo": "Relatório de Atendimentos",
        "descricao": "Atendimentos NYE/NYC concluídos no período, com resolução, tempo e monitor.",
    },
    "disparos": {
        "titulo": "Relatório de Disparos",
        "descricao": "Clientes com disparos de zona no período — uma linha por cliente.",
    },
    "disparos_geral": {
        "titulo": "Disparos Geral (fim de semana)",
        "descricao": "TODOS os disparos por cliente, classificados e agrupados em 3 abas (Villefort / Super Nosso / Base).",
    },
}

# módulos que geram planilha de 3 abas (prévia por grupo)
MODULOS_MULTIGRUPO = {"disparos_geral"}


def _agora() -> datetime:
    return datetime.now(FUSO_HORARIO)


def _parse_data(valor: str) -> datetime:
    return datetime.strptime(valor, "%Y-%m-%d").replace(tzinfo=FUSO_HORARIO)


def _parse_data_hora(valor: str) -> datetime:
    return datetime.strptime(valor, "%Y-%m-%dT%H:%M").replace(tzinfo=FUSO_HORARIO)


def _periodo_fim_de_semana(agora: datetime) -> tuple[datetime, datetime]:
    """Sexta 18h → segunda 08h da semana (mesma lógica do script). No
    próprio sábado/sexta, usa o fim de semana anterior (ainda em curso)."""
    dias = (agora.weekday() - 5) % 7 or 7  # sábado mais recente
    sabado = agora - timedelta(days=dias)
    inicio = (sabado - timedelta(days=1)).replace(hour=18, minute=0, second=0, microsecond=0)
    fim = (sabado + timedelta(days=2)).replace(hour=8, minute=0, second=0, microsecond=0)
    return inicio, fim


def _periodo_do_form(modulo: str) -> tuple[datetime, datetime] | None:
    """None = janela automática (só Disparos usa)."""
    preset = request.form.get("periodo", "ontem")
    agora = _agora()
    hoje = datetime.combine(agora.date(), time.min, tzinfo=FUSO_HORARIO)

    if preset == "auto":
        return None
    if preset == "fim_de_semana":
        return _periodo_fim_de_semana(agora)
    if preset == "ontem":
        inicio = hoje - timedelta(days=1)
        return inicio, hoje - timedelta(seconds=1)
    if preset == "7dias":
        return hoje - timedelta(days=7), agora
    # manual — Disparos e Disparos Geral usam hora/minuto; Atendimentos, dia inteiro
    if modulo in ("disparos", "disparos_geral"):
        inicio = _parse_data_hora(request.form["inicio"])
        fim = _parse_data_hora(request.form["fim"])
    else:
        inicio = _parse_data(request.form["inicio"])
        fim = _parse_data(request.form["fim"]) + timedelta(hours=23, minutes=59, seconds=59)
    if fim < inicio:
        raise ValueError("Período inválido: fim antes do início.")
    return inicio, fim


def _linhas_da_aba(aba) -> tuple[list[str], list[tuple]]:
    linhas = [
        tuple(c if c is not None else "" for c in linha)
        for linha in aba.iter_rows(values_only=True)
    ]
    if not linhas:
        return [], []
    return list(linhas[0]), linhas[1:]


def _ler_previa(run: ReportRun) -> tuple[list[str], list[tuple]]:
    """Lê a aba principal do .xlsx arquivado — a prévia é, por construção,
    idêntica ao Excel baixado."""
    if not run.file_path or not Path(run.file_path).exists():
        return [], []
    wb = load_workbook(run.file_path, read_only=True)
    cabecalho, linhas = _linhas_da_aba(wb.worksheets[0])
    wb.close()
    return cabecalho, linhas


def _ler_previa_grupos(run: ReportRun) -> tuple[list[str], list[dict]]:
    """Lê todas as abas (um grupo por aba) do .xlsx de Disparos Geral."""
    if not run.file_path or not Path(run.file_path).exists():
        return [], []
    wb = load_workbook(run.file_path, read_only=True)
    cabecalho: list[str] = []
    grupos: list[dict] = []
    for aba in wb.worksheets:
        cab, linhas = _linhas_da_aba(aba)
        if cab and not cabecalho:
            cabecalho = cab
        grupos.append({"nome": aba.title, "linhas": linhas})
    wb.close()
    return cabecalho, grupos


@bp.route("/<modulo>")
@login_required
def pagina(modulo: str):
    if modulo not in MODULOS:
        abort(404)

    historico = (
        ReportRun.query.filter_by(module=modulo)
        .order_by(ReportRun.generated_at.desc())
        .limit(HISTORICO_LIMITE)
        .all()
    )
    ultimo_ok = next((r for r in historico if r.status == "success"), None)

    multigrupo = modulo in MODULOS_MULTIGRUPO
    cabecalho: list[str] = []
    linhas: list[tuple] = []
    grupos: list[dict] = []
    if ultimo_ok is not None:
        if multigrupo:
            cabecalho, grupos = _ler_previa_grupos(ultimo_ok)
        else:
            cabecalho, linhas = _ler_previa(ultimo_ok)

    pagina_atual = max(request.args.get("pagina", 1, type=int), 1)
    total_paginas = max((len(linhas) + LINHAS_POR_PAGINA - 1) // LINHAS_POR_PAGINA, 1)
    inicio = (pagina_atual - 1) * LINHAS_POR_PAGINA
    linhas_pagina = linhas[inicio : inicio + LINHAS_POR_PAGINA]

    janela_auto = None
    if modulo == "disparos":
        janela_auto = report_service.janela_disparos(agora=_agora())
    fim_de_semana = _periodo_fim_de_semana(_agora()) if modulo == "disparos_geral" else None

    return render_template(
        "reports/pagina.html",
        modulo=modulo,
        info=MODULOS[modulo],
        multigrupo=multigrupo,
        historico=historico,
        ultimo_ok=ultimo_ok,
        cabecalho=cabecalho,
        linhas=linhas_pagina,
        grupos=grupos,
        total_linhas=len(linhas),
        pagina_atual=pagina_atual,
        total_paginas=total_paginas,
        janela_auto=janela_auto,
        fim_de_semana=fim_de_semana,
    )


@bp.route("/<modulo>/gerar", methods=["POST"])
@login_required
def gerar(modulo: str):
    if modulo not in MODULOS:
        abort(404)

    try:
        periodo = _periodo_do_form(modulo)
    except (ValueError, KeyError):
        flash("Período inválido — confira as datas.", "warning")
        return redirect(url_for("reports.pagina", modulo=modulo))

    try:
        if modulo == "atendimentos":
            if periodo is None:
                flash("Escolha um período para o relatório de atendimentos.", "warning")
                return redirect(url_for("reports.pagina", modulo=modulo))
            run = report_service.gerar_atendimentos(
                config=current_app.config,
                desde=periodo[0],
                hasta=periodo[1],
                user_id=current_user.id,
            )
        elif modulo == "disparos_geral":
            if periodo is None:
                flash("Escolha um período para o relatório de Disparos Geral.", "warning")
                return redirect(url_for("reports.pagina", modulo=modulo))
            run = report_service.gerar_disparos_geral(
                config=current_app.config,
                desde=periodo[0],
                hasta=periodo[1],
                user_id=current_user.id,
            )
        else:
            desde, hasta = periodo if periodo is not None else (None, None)
            run = report_service.gerar_disparos(
                config=current_app.config, user_id=current_user.id, desde=desde, hasta=hasta
            )
    except report_service.RelatorioEmAndamentoError:
        audit_service.registrar(
            action="report_generated",
            result="failure",
            user=current_user,
            details={"module": modulo, "motivo": "geracao_em_andamento"},
        )
        db.session.commit()
        flash("Já existe uma geração deste relatório em andamento. Aguarde terminar.", "warning")
        return redirect(url_for("reports.pagina", modulo=modulo))

    audit_service.registrar(
        action="report_generated",
        result="success" if run.status == "success" else "failure",
        user=current_user,
        details={
            "module": modulo,
            "run_id": run.id,
            "period_start": run.period_start.isoformat(),
            "period_end": run.period_end.isoformat(),
            "row_count": run.row_count,
            "extra_counts": run.extra_counts,
            "error": run.error_message,
        },
    )
    db.session.commit()

    if run.status == "success":
        flash("Relatório gerado com sucesso.", "info")
    else:
        flash(f"A geração terminou com erro: {run.error_message}", "error")
    return redirect(url_for("reports.pagina", modulo=modulo))


@bp.route("/download/<int:run_id>")
@login_required
def download(run_id: int):
    run = db.session.get(ReportRun, run_id)
    if run is None or not run.file_path or not Path(run.file_path).exists():
        abort(404)
    return send_file(
        run.file_path,
        as_attachment=True,
        download_name=Path(run.file_path).name,
    )
