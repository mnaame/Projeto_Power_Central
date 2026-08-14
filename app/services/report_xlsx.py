from __future__ import annotations

from pathlib import Path
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Estilo validado com a operação (mesmo das planilhas manuais que estes
# relatórios substituem): cabeçalho verde #21A366, texto branco.
COR_CABECALHO = "21A366"

COLUNAS_ATENDIMENTOS = (
    "DATA EVENTO",
    "CONTA",
    "CLIENTE",
    "EVENTO",
    "SITUAÇÃO",
    "TEMPO DE ATENDIMENTO",
    "MONITOR",
)
COLUNAS_DESCARTADOS = ("DATA", "CONTA", "CLIENTE", "EVENTO", "MOTIVO")
COLUNAS_DISPAROS_GERAL = (
    "CLIENTE",
    "QUANTIDADE DE DISPARO",
    "OCORRENCIA",
    "TEMP/CONCLUSÃO",
    "TEMPO PARA LIGAR PARA O CLIENTE",
    "ZONA",
)
COLUNAS_DISPAROS = (
    "CLIENTE",
    "QUANTIDADE DE DISPARO",
    "HORÁRIOS DOS DISPAROS",
    "OCORRENCIA",
    "TEMP/CONCLUSÃO",
    "TEMPO PARA LIGAR PARA O CLIENTE",
    "ZONA",
    "CONCLUSÃO MONITORAMENTO DISPARO.",
)
COLUNAS_CENTRAL_CLIENTE = (
    "CLIENTE",
    "ID AUVO",
    "TELEFONE",
    "LOGIN",
    "LINK DE ACESSO",
    "CÓDIGO DO CONTATO",
    "STATUS",
)
COLUNAS_BI_INTERVENCOES = (
    "DATA DA VISITA",
    "CONTA",
    "LOJA",
    "TÉCNICO",
    "DISPAROS/DIA ANTES",
    "DISPAROS/DIA DEPOIS",
    "VARIAÇÃO %",
    "CLASSIFICAÇÃO",
    "JANELA PARCIAL",
    "OUTRA VISITA NA JANELA",
)
COLUNAS_BI_CRONICOS = (
    "CONTA",
    "LOJA",
    "Nº DE VISITAS",
    "ÚLTIMA CLASSIFICAÇÃO",
    "DISPAROS/DIA ATUAL",
)

_LARGURAS = {
    "DATA EVENTO": 18,
    "DATA": 18,
    "CONTA": 10,
    "CLIENTE": 40,
    "EVENTO": 10,
    "SITUAÇÃO": 60,
    "TEMPO DE ATENDIMENTO": 22,
    "MONITOR": 20,
    "MOTIVO": 50,
    "QUANTIDADE DE DISPARO": 22,
    "HORÁRIOS DOS DISPAROS": 20,
    "OCORRENCIA": 26,
    "TEMP/CONCLUSÃO": 18,
    "TEMPO PARA LIGAR PARA O CLIENTE": 30,
    "ZONA": 40,
    "CONCLUSÃO MONITORAMENTO DISPARO.": 40,
    "DATA DA VISITA": 18,
    "LOJA": 40,
    "TÉCNICO": 22,
    "DISPAROS/DIA ANTES": 20,
    "DISPAROS/DIA DEPOIS": 20,
    "VARIAÇÃO %": 14,
    "CLASSIFICAÇÃO": 16,
    "JANELA PARCIAL": 16,
    "OUTRA VISITA NA JANELA": 20,
    "Nº DE VISITAS": 16,
    "ÚLTIMA CLASSIFICAÇÃO": 20,
    "DISPAROS/DIA ATUAL": 20,
    "ID AUVO": 12,
    "TELEFONE": 18,
    "LOGIN": 22,
    "LINK DE ACESSO": 60,
    "CÓDIGO DO CONTATO": 16,
    "STATUS": 14,
}
_COLUNAS_COM_QUEBRA = {
    "SITUAÇÃO",
    "ZONA",
    "MOTIVO",
    "CONCLUSÃO MONITORAMENTO DISPARO.",
    "HORÁRIOS DOS DISPAROS",
}


def _montar_aba(ws, colunas: Sequence[str], linhas: Sequence[Sequence[object]]) -> None:
    fonte_cabecalho = Font(color="FFFFFF", bold=True)
    fundo_cabecalho = PatternFill(start_color=COR_CABECALHO, end_color=COR_CABECALHO, fill_type="solid")

    for indice, titulo in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=indice, value=titulo)
        celula.font = fonte_cabecalho
        celula.fill = fundo_cabecalho
        ws.column_dimensions[get_column_letter(indice)].width = _LARGURAS.get(titulo, 18)

    for numero_linha, linha in enumerate(linhas, start=2):
        for indice, valor in enumerate(linha, start=1):
            celula = ws.cell(row=numero_linha, column=indice, value=valor)
            if colunas[indice - 1] in _COLUNAS_COM_QUEBRA:
                celula.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                celula.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ultima_coluna = get_column_letter(len(colunas))
    ws.auto_filter.ref = f"A1:{ultima_coluna}{max(len(linhas) + 1, 1)}"


def gerar_xlsx_atendimentos(
    caminho: Path,
    *,
    incluidos: Sequence[Sequence[object]],
    descartados: Sequence[Sequence[object]],
) -> None:
    """Aba principal (A.4) + aba DESCARTADOS (prestação de contas)."""
    wb = Workbook()

    ws = wb.active
    ws.title = "ATENDIMENTOS"
    _montar_aba(ws, COLUNAS_ATENDIMENTOS, incluidos)

    ws_descartados = wb.create_sheet("DESCARTADOS")
    _montar_aba(ws_descartados, COLUNAS_DESCARTADOS, descartados)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def gerar_xlsx_disparos(caminho: Path, *, linhas: Sequence[Sequence[object]]) -> None:
    """Uma linha por cliente (B.4)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "DISPAROS"
    _montar_aba(ws, COLUNAS_DISPAROS, linhas)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def gerar_xlsx_bi_intervencoes(caminho: Path, *, linhas: Sequence[Sequence[object]]) -> None:
    """Uma linha por intervenção (ordem concluída × janela antes/depois)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "INTERVENCOES"
    _montar_aba(ws, COLUNAS_BI_INTERVENCOES, linhas)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def gerar_xlsx_bi_cronicos(caminho: Path, *, linhas: Sequence[Sequence[object]]) -> None:
    """Uma linha por cliente crônico (visitas repetidas sem melhora real)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CRONICOS"
    _montar_aba(ws, COLUNAS_BI_CRONICOS, linhas)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def gerar_xlsx_disparos_geral(
    caminho: Path,
    *,
    por_grupo: dict[str, Sequence[Sequence[object]]],
    grupos: Sequence[str],
) -> None:
    """Uma planilha com uma aba por grupo (Villefort / Super Nosso / Base),
    na ordem dada. Mesmas colunas em todas."""
    wb = Workbook()
    wb.remove(wb.active)
    for grupo in grupos:
        ws = wb.create_sheet(grupo)
        _montar_aba(ws, COLUNAS_DISPAROS_GERAL, por_grupo.get(grupo, []))

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)


def gerar_xlsx_central_cliente(caminho: Path, *, linhas: Sequence[Sequence[object]]) -> None:
    """Uma linha por item do lote (Links da Central do Cliente) — documento
    sensível (login/link de acesso sem senha), mesmo cuidado do Cofre."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CENTRAL_CLIENTE"
    _montar_aba(ws, COLUNAS_CENTRAL_CLIENTE, linhas)

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
