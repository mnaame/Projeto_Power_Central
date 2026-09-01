"""Relatório do Técnico do Dia — envelopa o motor validado do
`relatorio_tecnico.py`: match de técnico na tarefa da Auvo, nome de
arquivo por loja, e a conversão do HTML nativo da plataforma (o export
"Excel" real é HTML) para um `.xlsx` de verdade preservando as cores por
tipo de evento.

Fica em `domain/` por ser lógica pura (sem rede) — a única parte "impura"
é a montagem do Workbook em memória; salvar em disco é responsabilidade
de quem chama (`services/tecnico_service.py`).
"""

from __future__ import annotations

import html as _html
import re
import unicodedata
from dataclasses import dataclass
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Campos tentados em ordem — o primeiro não-vazio identifica o técnico da
# tarefa (nomes variam conforme a versão da API da Auvo).
_CAMPOS_TECNICO = ("userToName", "idUserToName", "userTo", "responsavel", "idUserTo")

# Mesma ideia para o cliente Auvo da tarefa (id para o de-para reverso,
# nome só para mostrar quando não há vínculo) e para o horário — nenhum
# desses três foi validado contra uma tarefa real de agenda (só os campos
# de status o foram, contra a tarefa 77330829); por isso são vários
# candidatos e nunca um erro — pior caso é a coluna ficar vazia na tela.
_CAMPOS_ID_CLIENTE = ("customerId", "idCustomer")
_CAMPOS_NOME_CLIENTE = ("customerDescription", "customerName")
_CAMPOS_HORARIO = ("taskDate", "startDatetime", "startDate", "scheduledDate", "date")

_MARCADOR_ERRO_1 = "no se encontr"
_MARCADOR_ERRO_2 = "regularizar la situaci"

_LARGURA_COLUNA_PADRAO = 22
COR_CABECALHO = "21A366"
_TEXTO_CABECALHO = "data e hora do evento"

TAM_MAX_NOME_ARQUIVO = 40


def _texto(valor: object) -> str:
    return str(valor).strip() if valor is not None else ""


def _sem_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )


def nome_tecnico_da_tarefa(tarefa: Mapping[str, object]) -> str:
    """Primeiro campo não-vazio entre os candidatos conhecidos."""
    for campo in _CAMPOS_TECNICO:
        valor = _texto(tarefa.get(campo))
        if valor:
            return valor
    return ""


def tecnico_corresponde(tarefa: Mapping[str, object], alvo: str) -> bool:
    """Substring, sem acento, case-insensitive. Alvo vazio = "todos"."""
    alvo = _texto(alvo)
    if not alvo:
        return True
    nome = _sem_acentos(nome_tecnico_da_tarefa(tarefa)).lower()
    return _sem_acentos(alvo).lower() in nome


def id_cliente_da_tarefa(tarefa: Mapping[str, object]) -> int | None:
    """Id do cliente Auvo da tarefa — chave do de-para reverso."""
    for campo in _CAMPOS_ID_CLIENTE:
        valor = tarefa.get(campo)
        if valor is not None:
            try:
                return int(valor)
            except (TypeError, ValueError):
                continue
    return None


def nome_cliente_da_tarefa(tarefa: Mapping[str, object]) -> str:
    """Nome do cliente Auvo, só para exibir quando a loja não tem
    vínculo no de-para (com vínculo, o nome que conta é o da PowerCentral)."""
    for campo in _CAMPOS_NOME_CLIENTE:
        valor = _texto(tarefa.get(campo))
        if valor:
            return valor
    id_cliente = id_cliente_da_tarefa(tarefa)
    return str(id_cliente) if id_cliente is not None else ""


def horario_da_tarefa(tarefa: Mapping[str, object]) -> str:
    """Só para exibir na tela — nunca usado em regra de negócio, então um
    campo não encontrado não é erro, fica em branco."""
    for campo in _CAMPOS_HORARIO:
        valor = _texto(tarefa.get(campo))
        if valor:
            return valor
    return ""


def mapa_contas(contas_brutas: Sequence[Mapping[str, object]]) -> dict[str, tuple[str, str]]:
    """Número normalizado (sem zeros à esquerda) -> (cue_iid, nome), a
    partir da lista crua de `listar_todas_contas`. Mesma normalização de
    conta usada no resto do sistema (`auvo_service.normalizar_conta`)."""
    mapa: dict[str, tuple[str, str]] = {}
    for conta in contas_brutas:
        numero = _texto(conta.get("cue_ncuenta")).lstrip("0") or "0"
        cue_iid = conta.get("cue_iid") or conta.get("Id")
        if cue_iid is None:
            continue
        mapa[numero] = (str(cue_iid), _texto(conta.get("cue_cnombre")))
    return mapa


def sanitizar_nome_arquivo(nome: str, *, tamanho_maximo: int = TAM_MAX_NOME_ARQUIVO) -> str:
    """Só alfanumérico, espaço, hífen e underline — trunca. Mesma regra
    do motor validado, para os nomes dos arquivos baterem com o que a
    operação já está acostumada a ver."""
    limpo = "".join(c if c.isalnum() or c in " -_" else "_" for c in (nome or ""))
    return limpo[:tamanho_maximo].strip()


def nome_arquivo_loja(numero_conta: str, nome_cliente: str, *, extensao: str = "xls") -> str:
    """`<conta 4 dígitos>_<nome sanitizado>.<extensao>` — mesmo padrão do
    motor validado (CuentaNumero também é enviado com 4 dígitos)."""
    numero = (numero_conta or "0").zfill(4)
    return f"{numero}_{sanitizar_nome_arquivo(nome_cliente)}.{extensao}"


def exportacao_recusada(conteudo: bytes) -> bool:
    """True quando o HTML devolvido é a página de erro de permissão da
    PowerCentral, não o histórico de verdade."""
    texto = conteudo.decode("utf-8", "replace").lower()
    return _MARCADOR_ERRO_1 in texto or _MARCADOR_ERRO_2 in texto


# Extração das linhas do HTML nativo do export — compartilhada entre a
# conversão colorida e a contagem de eventos, para as duas lerem o
# arquivo do mesmo jeito.
_RE_LINHA = re.compile(r"<tr[^>]*>(.*?)</tr>", re.S | re.I)
_RE_CELULA = re.compile(r"<t[dh]([^>]*)>(.*?)</t[dh]>", re.S | re.I)


def _texto_da_celula(conteudo: str) -> str:
    texto = re.sub(r"<[^>]+>", "", conteudo)
    return _html.unescape(texto).replace("\xa0", " ").strip()


@dataclass(frozen=True)
class CelulaExport:
    texto: str
    cor_fundo: str | None
    cor_texto: str | None


def _para_texto(conteudo: bytes | str) -> str:
    return conteudo.decode("utf-8", "replace") if isinstance(conteudo, bytes) else conteudo


def linha_e_cabecalho(linha: Sequence[CelulaExport]) -> bool:
    return _TEXTO_CABECALHO in " ".join(c.texto for c in linha).lower()


def linhas_do_export(conteudo: bytes | str) -> list[list[CelulaExport]]:
    """Linhas úteis do HTML nativo do export (sem as vazias), com as cores
    de cada célula. Base única do `.xlsx`, do `.pdf` e da contagem — as
    três precisam ler o arquivo exatamente do mesmo jeito."""
    linhas: list[list[CelulaExport]] = []
    for linha_html in _RE_LINHA.findall(_para_texto(conteudo)):
        celulas = [
            CelulaExport(
                texto=_texto_da_celula(conteudo_cel),
                cor_fundo=_cor_do_estilo(_estilo(atributos), "background-color"),
                cor_texto=_cor_do_estilo(_estilo(atributos), "color"),
            )
            for atributos, conteudo_cel in _RE_CELULA.findall(linha_html)
        ]
        if any(c.texto for c in celulas):
            linhas.append(celulas)
    return linhas


def contar_eventos_do_export(conteudo: bytes | str) -> int:
    """Quantos eventos o export tem, fora o cabeçalho e as linhas vazias.

    Conta a partir do próprio arquivo porque `buscar_historico` NÃO filtra
    por conta — usá-lo para o resumo de uma loja daria o número da base
    inteira (e uma consulta enorme por cima)."""
    return sum(1 for linha in linhas_do_export(conteudo) if not linha_e_cabecalho(linha))


def _estilo(atributos: str) -> str:
    encontrado = re.search(r'style\s*=\s*"([^"]*)"', atributos or "")
    return encontrado.group(1) if encontrado else ""


def _cor_do_estilo(estilo: str, propriedade: str) -> str | None:
    encontrado = re.search(propriedade + r"\s*:\s*#([0-9a-fA-F]{6})", estilo or "")
    return encontrado.group(1).upper() if encontrado else None


def montar_workbook_colorido(html: str) -> Workbook:
    """Converte o HTML nativo do export num `.xlsx` de verdade,
    preservando as cores de fundo/texto de cada evento (mesma lógica do
    motor validado). Devolve o Workbook em memória — salvar é por conta
    de quem chama."""
    workbook = Workbook()
    planilha = workbook.active
    planilha.title = "Historico"
    cabecalho_marcado = False

    for numero, linha in enumerate(linhas_do_export(html), start=1):
        eh_cabecalho = linha_e_cabecalho(linha)
        for coluna, celula_export in enumerate(linha, start=1):
            celula = planilha.cell(row=numero, column=coluna, value=celula_export.texto)
            if eh_cabecalho:
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill("solid", fgColor=COR_CABECALHO)
            else:
                if celula_export.cor_fundo and celula_export.cor_fundo != "TRANSPARENT":
                    celula.fill = PatternFill("solid", fgColor="FF" + celula_export.cor_fundo)
                if celula_export.cor_texto:
                    celula.font = Font(color="FF" + celula_export.cor_texto)
            celula.alignment = Alignment(vertical="top", wrap_text=not eh_cabecalho)

        if eh_cabecalho and not cabecalho_marcado:
            planilha.freeze_panes = f"A{numero + 1}"
            cabecalho_marcado = True
            for coluna in range(1, len(linha) + 1):
                planilha.column_dimensions[get_column_letter(coluna)].width = _LARGURA_COLUNA_PADRAO

    return workbook


def montar_pdf_colorido(conteudo: bytes | str, *, titulo: str = "") -> bytes:
    """Mesmo conteúdo do `.xlsx`, em PDF — abre no celular sem app de
    planilha. Preserva as cores por tipo de evento, repete o cabeçalho em
    toda página e usa paisagem (a tabela é larga).

    Lê pelo `linhas_do_export`, então PDF e XLS mostram exatamente as
    mesmas linhas: se um divergisse do outro, o técnico não saberia em
    qual acreditar."""
    from io import BytesIO

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    linhas = linhas_do_export(conteudo)
    estilo_celula = ParagraphStyle("celula", fontName="Helvetica", fontSize=7, leading=9)
    estilo_cabecalho = ParagraphStyle(
        "cabecalho", fontName="Helvetica-Bold", fontSize=7, leading=9,
        textColor=colors.white,
    )

    dados: list[list] = []
    comandos: list[tuple] = [
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0B0B0")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
    ]
    indice_cabecalho = None
    for numero, linha in enumerate(linhas):
        eh_cabecalho = linha_e_cabecalho(linha)
        if eh_cabecalho and indice_cabecalho is None:
            indice_cabecalho = numero
        estilo = estilo_cabecalho if eh_cabecalho else estilo_celula
        dados.append([Paragraph(_html.escape(c.texto), estilo) for c in linha])

        if eh_cabecalho:
            comandos.append(
                ("BACKGROUND", (0, numero), (-1, numero), colors.HexColor("#" + COR_CABECALHO))
            )
            continue
        for coluna, celula in enumerate(linha):
            if celula.cor_fundo and celula.cor_fundo != "TRANSPARENT":
                comandos.append(
                    (
                        "BACKGROUND",
                        (coluna, numero),
                        (coluna, numero),
                        colors.HexColor("#" + celula.cor_fundo),
                    )
                )
            if celula.cor_texto:
                comandos.append(
                    (
                        "TEXTCOLOR",
                        (coluna, numero),
                        (coluna, numero),
                        colors.HexColor("#" + celula.cor_texto),
                    )
                )

    saida = BytesIO()
    documento = SimpleDocTemplate(
        saida,
        pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
        title=titulo or "Histórico",
    )
    elementos = []
    if titulo:
        elementos.append(
            Paragraph(
                _html.escape(titulo),
                ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=11, leading=14),
            )
        )
        elementos.append(Spacer(1, 4 * mm))

    if dados:
        # `repeatRows` só faz sentido quando o cabeçalho é a 1ª linha —
        # se vier depois (ou não vier), não repete nada em vez de repetir
        # uma linha de evento no topo de cada página.
        repetir = 1 if indice_cabecalho == 0 else 0
        tabela = Table(dados, repeatRows=repetir, hAlign="LEFT")
        tabela.setStyle(TableStyle(comandos))
        elementos.append(tabela)
    else:
        elementos.append(Paragraph("Nenhum evento no período.", estilo_celula))

    documento.build(elementos)
    return saida.getvalue()
