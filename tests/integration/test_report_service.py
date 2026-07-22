from datetime import datetime, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from app.extensions import db
from app.models.report import ReportRun
from app.services import report_service, retention_service, settings_service

FUSO = ZoneInfo("America/Sao_Paulo")
DESDE = datetime(2026, 7, 18, 0, 0, 0, tzinfo=FUSO)
HASTA = datetime(2026, 7, 18, 23, 59, 59, tzinfo=FUSO)


def _fmt(dt):
    return dt.astimezone(FUSO).strftime("%m/%d/%Y %I:%M:%S %p")


class FakeSoftGuardClient:
    """Devolve eventos do histórico e timelines por id, e registra as
    chamadas para as asserções de parâmetros."""

    def __init__(self, eventos=None, timelines=None):
        self.eventos = eventos or []
        self.timelines = timelines or {}
        self.chamadas_historico = []
        self.chamadas_timeline = []

    def buscar_historico(self, *, codigos_alarme, desde, hasta, **kwargs):
        self.chamadas_historico.append(
            {"codigos": tuple(codigos_alarme), "desde": desde, "hasta": hasta}
        )
        return self.eventos

    def buscar_timeline(self, id_evento, **kwargs):
        self.chamadas_timeline.append(id_evento)
        return self.timelines.get(str(id_evento), [])


def _evento_nye(rec_iid="500", quando=None, conta="0004", cliente="VILLEFORT TROPICAL"):
    return {
        "rec_iid": rec_iid,
        "rec_calarma": "NYE",
        "rec_tfechahora": _fmt(quando or DESDE + timedelta(hours=21, minutes=30)),
        "cue_ncuenta": conta,
        "cue_cnombre": cliente,
        "rec_iidcuenta": "9385",
    }


def _timeline_fechada(comentario="Cliente vai ativar depois"):
    return [
        {
            "etl_tFechaHora": "7/18/2026 9:30:00 PM",
            "etl_cAccion": "Inicio",
            "etl_cObservacion": "Evento recebido na Central",
            "etl_iAccionCode": "",
            "ope_cnombre": "",
        },
        {
            "etl_tFechaHora": "7/18/2026 9:35:00 PM",
            "etl_cAccion": "IngresoComentarios",
            "etl_cObservacion": comentario,
            "etl_iAccionCode": "",
            "ope_cnombre": "MARIA",
        },
        {
            "etl_tFechaHora": "7/18/2026 9:41:05 PM",
            "etl_cAccion": "Procesar",
            "etl_cObservacion": "Evento processado",
            "etl_iAccionCode": "122",
            "ope_cnombre": "MARIA",
        },
    ]


# ---------- Atendimentos ----------


def test_gerar_atendimentos_fim_a_fim(app):
    client = FakeSoftGuardClient(
        eventos=[
            _evento_nye(rec_iid="500"),
            _evento_nye(rec_iid="501", conta="0141", cliente="VILLEFORT FONTE GRANDE"),
        ],
        timelines={
            "500": _timeline_fechada("Responsável avisado, vai ativar"),
            "501": _timeline_fechada("Sistema ativado pelo responsável"),  # descartado
        },
    )

    run = report_service.gerar_atendimentos(
        config=app.config, desde=DESDE, hasta=HASTA, user_id=None, softguard_client=client
    )

    assert run.status == "success"
    assert run.row_count == 1
    assert run.extra_counts == {"descartados": 1, "total_eventos": 2}
    assert client.chamadas_historico[0]["codigos"] == ("NYE", "NYC")

    wb = load_workbook(run.file_path)
    aba = wb["ATENDIMENTOS"]
    cabecalho = [c.value for c in aba[1]]
    assert cabecalho == [
        "DATA EVENTO", "CONTA", "CLIENTE", "EVENTO", "SITUAÇÃO",
        "TEMPO DE ATENDIMENTO", "MONITOR",
    ]
    linha = [c.value for c in aba[2]]
    assert linha[1] == "0004"
    assert linha[3] == "NYE"
    assert "SAB:" in linha[4]
    assert linha[5] == "00H11M05S"
    assert linha[6] == "MARIA"
    assert aba.freeze_panes == "A2"
    assert aba.auto_filter.ref is not None

    descartados = wb["DESCARTADOS"]
    linha_descartada = [c.value for c in descartados[2]]
    assert linha_descartada[1] == "0141"
    assert "armou" in linha_descartada[4]

    # estilo do cabeçalho (aceite visual do Excel)
    celula = aba["A1"]
    assert celula.fill.start_color.rgb.endswith("21A366")
    assert celula.font.bold is True


def test_gerar_atendimentos_erro_do_portal_vira_status_error(app):
    class ClienteQuebrado:
        def buscar_historico(self, **kwargs):
            raise RuntimeError("portal fora do ar")

    run = report_service.gerar_atendimentos(
        config=app.config, desde=DESDE, hasta=HASTA, user_id=None,
        softguard_client=ClienteQuebrado(),
    )
    assert run.status == "error"
    assert "portal fora do ar" in run.error_message


def test_lock_impede_geracao_concorrente(app):
    lock = report_service._locks["atendimentos"]
    assert lock.acquire(blocking=False)
    try:
        with pytest.raises(report_service.RelatorioEmAndamentoError):
            report_service.gerar_atendimentos(
                config=app.config, desde=DESDE, hasta=HASTA, user_id=None,
                softguard_client=FakeSoftGuardClient(),
            )
    finally:
        lock.release()


# ---------- Disparos ----------


def _evento_bur(rec_iid, quando, conta="9385", cliente="CLIENTE X", operador="7",
                zona="(28) IVP ENTRADA RM"):
    return {
        "rec_iid": rec_iid,
        "rec_iidcuenta": conta,
        "cue_cnombre": cliente,
        "rec_calarma": "BUR",
        "rec_tfechahora": _fmt(quando),
        "_zon_cdescripcion": zona,
        "rec_ioperador": operador,
    }


def test_gerar_disparos_fim_a_fim_com_janela_movel(app):
    meio = DESDE + timedelta(hours=12)
    client = FakeSoftGuardClient(
        eventos=[
            _evento_bur("700", meio),
            _evento_bur("701", meio + timedelta(minutes=10)),
        ],
        timelines={"701": _timeline_fechada("Contato feito")},
    )

    run = report_service.gerar_disparos(
        config=app.config, user_id=None, desde=DESDE, hasta=HASTA, softguard_client=client
    )

    assert run.status == "success"
    assert run.row_count == 1
    assert run.extra_counts == {"total_disparos": 2, "clientes": 1}

    # consulta leva a folga de 6 min nas duas pontas
    chamada = client.chamadas_historico[0]
    assert chamada["desde"] == (DESDE - timedelta(minutes=6)).astimezone(FUSO)
    assert chamada["hasta"] == (HASTA + timedelta(minutes=6)).astimezone(FUSO)
    assert set(chamada["codigos"]) == {"BUR", "CLO", "CLV", "ROP", "OPN", "OPV", "RCL"}

    wb = load_workbook(run.file_path)
    aba = wb["DISPAROS"]
    linha = [c.value for c in aba[2]]
    assert linha[0] == "CLIENTE X"
    assert linha[1] == "2x"
    assert linha[2] == "ALEATORIO"
    assert linha[3] == "00H11M05S"  # tempo do disparo atendido mais recente
    assert linha[4] == "X"
    assert linha[5] == "(28) IVP ENTRADA RM"
    assert linha[6] is None or linha[6] == ""


def test_janela_movel_encadeia_com_o_relatorio_anterior(app):
    client = FakeSoftGuardClient()

    primeiro = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client
    )
    assert primeiro.status == "success"
    # primeira execução: 24h (padrão) para trás
    duracao = primeiro.period_end - primeiro.period_start
    assert abs(duracao - timedelta(hours=24)) < timedelta(minutes=1)

    segundo = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client
    )
    assert segundo.period_start == primeiro.period_end


def test_janela_manual_nao_altera_encadeamento_automatico(app):
    client = FakeSoftGuardClient()

    report_service.gerar_disparos(config=app.config, user_id=None, softguard_client=client)
    manual = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client,
        desde=DESDE - timedelta(days=30), hasta=DESDE - timedelta(days=29),
    )
    assert manual.status == "success"

    # o próximo automático parte do run de period_end mais recente
    proximo = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client
    )
    assert proximo.period_start > DESDE  # não voltou 30 dias


def test_janela_manual_com_fim_no_futuro_nao_trava_encadeamento_automatico(app):
    client = FakeSoftGuardClient()
    agora = datetime.now(timezone.utc)

    # relatório manual com fim no futuro (ex.: "hoje até 23:59" gerado de manhã)
    manual = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client,
        desde=agora - timedelta(hours=2), hasta=agora + timedelta(hours=10),
    )
    assert manual.status == "success"

    # o próximo automático não pode ficar "preso" no fim futuro do manual
    proximo = report_service.gerar_disparos(
        config=app.config, user_id=None, softguard_client=client
    )
    assert proximo.status == "success"
    assert proximo.period_start <= proximo.period_end


def test_retencao_remove_relatorio_antigo_e_arquivo(app, tmp_path):
    arquivo = tmp_path / "antigo.xlsx"
    arquivo.write_bytes(b"conteudo")

    agora = datetime.now(timezone.utc)
    run = ReportRun(
        module="disparos",
        generated_at=agora - timedelta(days=100),
        period_start=agora - timedelta(days=101),
        period_end=agora - timedelta(days=100),
        status="success",
        file_path=str(arquivo),
    )
    db.session.add(run)
    db.session.commit()

    resultado = retention_service.limpar_dados_antigos(retention_days=90, agora=agora)

    assert resultado["relatorios_removidos"] == 1
    assert not arquivo.exists()
    assert ReportRun.query.count() == 0
