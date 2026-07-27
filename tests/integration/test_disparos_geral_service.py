from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.models.report import ReportRun
from app.services import report_service

FUSO = ZoneInfo("America/Sao_Paulo")
# fim de semana: sexta 18h -> segunda 08h
DESDE = datetime(2026, 7, 24, 18, 0, 0, tzinfo=FUSO)
HASTA = datetime(2026, 7, 27, 8, 0, 0, tzinfo=FUSO)


def _fmt(dt):
    return dt.astimezone(FUSO).strftime("%m/%d/%Y %I:%M:%S %p")


def _evento(codigo, quando, *, conta, numero, nome, rec_iid="", zona="(28) IVP ENTRADA RM",
            operador="0"):
    return {
        "rec_iid": rec_iid,
        "rec_iidcuenta": conta,
        "cue_ncuenta": numero,
        "cue_cnombre": nome,
        "rec_calarma": codigo,
        "rec_tfechahora": _fmt(quando),
        "_zon_cdescripcion": zona,
        "rec_ioperador": operador,
    }


def _timeline_com_chamada():
    return [
        {"etl_tFechaHora": "7/25/2026 10:00:00 PM", "etl_cAccion": "Inicio",
         "etl_cObservacion": "Evento recebido", "etl_iAccionCode": "", "ope_cnombre": ""},
        {"etl_tFechaHora": "7/25/2026 10:09:00 PM", "etl_cAccion": "LlamadoTelefonico",
         "etl_cObservacion": "(*84 Chamada Atendida - Bem Sucedida)", "etl_iAccionCode": "",
         "ope_cnombre": ""},
        {"etl_tFechaHora": "7/25/2026 10:11:00 PM", "etl_cAccion": "Procesar",
         "etl_cObservacion": "processado", "etl_iAccionCode": "122", "ope_cnombre": "MARIA"},
    ]


class FakeSoftGuard:
    def __init__(self, eventos, timelines=None):
        self.eventos = eventos
        self.timelines = timelines or {}
        self.chamadas_historico = []

    def buscar_historico(self, *, codigos_alarme, desde, hasta, **kwargs):
        self.chamadas_historico.append({"codigos": tuple(codigos_alarme), "desde": desde, "hasta": hasta})
        return self.eventos

    def buscar_timeline(self, id_evento, **kwargs):
        return self.timelines.get(str(id_evento), [])


def test_gerar_disparos_geral_tres_abas_e_grupos(app):
    base = DESDE + timedelta(hours=4)
    eventos = [
        # Villefort: 1 disparo aleatório
        _evento("BUR", base, conta="1", numero="10", nome="VILLEFORT HM", rec_iid="v1"),
        # Super Nosso (APOIO): disparo após arme + atendido (tem timeline)
        _evento("CLO", base, conta="2", numero="20", nome="APOIO CURVELO"),
        _evento("BUR", base + timedelta(minutes=2), conta="2", numero="20",
                nome="APOIO CURVELO", rec_iid="s1", operador="7"),
        # Base: 2 disparos, um repetido (dedup)
        _evento("BUR", base, conta="3", numero="30", nome="ABC MERCADO", rec_iid="b1"),
        _evento("BUR", base, conta="3", numero="30", nome="ABC MERCADO", rec_iid="b1"),
    ]
    client = FakeSoftGuard(eventos, timelines={"s1": _timeline_com_chamada()})

    run = report_service.gerar_disparos_geral(
        config=app.config, desde=DESDE, hasta=HASTA, user_id=None, softguard_client=client
    )

    assert run.status == "success"
    assert run.row_count == 3
    assert run.extra_counts["por_grupo"] == {"Villefort": 1, "Super Nosso": 1, "Base": 1}
    # consulta com folga de 6 min e RCL na lista de códigos
    chamada = client.chamadas_historico[0]
    assert chamada["desde"] == (DESDE - timedelta(minutes=6)).astimezone(FUSO)
    assert set(chamada["codigos"]) == {"BUR", "CLO", "CLV", "ROP", "OPN", "OPV", "RCL"}

    wb = load_workbook(run.file_path)
    assert wb.sheetnames == ["Villefort", "Super Nosso", "Base"]

    # aba Super Nosso: APOIO, 1 disparo, após arme, com tempos do timeline
    aba = wb["Super Nosso"]
    cabecalho = [c.value for c in aba[1]]
    assert cabecalho == [
        "CLIENTE", "QUANTIDADE DE DISPARO", "OCORRENCIA", "TEMP/CONCLUSÃO",
        "TEMPO PARA LIGAR PARA O CLIENTE", "ZONA",
    ]
    linha = [c.value for c in aba[2]]
    assert linha[0] == "APOIO CURVELO"
    assert linha[1] == "1X"  # X maiúsculo
    assert linha[2] == "DISPARO APÓS ARME"
    assert linha[3] == "00H11M00S"  # conclusão (início 22:00 -> 22:11)
    assert linha[4] == "00H09M00S"  # ligação (início 22:00 -> chamada 22:09)

    # aba Base: dedup deixou 1 disparo; sem atendimento -> tempos vazios/X
    aba_base = wb["Base"]
    linha_base = [c.value for c in aba_base[2]]
    assert linha_base[0] == "ABC MERCADO"
    assert linha_base[1] == "1X"
    assert linha_base[4] == "X"  # sem ligação


def test_periodo_manual_registrado(app):
    client = FakeSoftGuard([])
    run = report_service.gerar_disparos_geral(
        config=app.config, desde=DESDE, hasta=HASTA, user_id=None, softguard_client=client
    )
    assert run.status == "success"
    assert run.period_start == DESDE
    assert run.period_end == HASTA
    assert run.row_count == 0
